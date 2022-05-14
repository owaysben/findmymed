# library required
from openpyxl import load_workbook
import mysql.connector

DB_USER = "root"
DB_PASS = ""
DB_HOST = "127.0.0.1"
DB_NAME = "findmymed"

# parsing data
try:
	file = load_workbook('bd.xlsx') # put here the name of the excel file if it exist in same directory of this script file else give the full path of excel file"
	sheet = file.active
	print("reading data from "+sheet.title+" ...")

	try:

		db = mysql.connector.connect(user=DB_USER,password=DB_PASS, host=DB_HOST, database=DB_NAME)

		pointer = db.cursor()

		sql = ("INSERT INTO products " "(name, forme, voie) " "VALUES (%s, %s, %s)")

		for row in sheet.rows:
			data = list(())
			for col in range(0,sheet.max_column):
				if ',' in  row[col].value:
					data.append(''.join(row[col].value.split(',')[:-1]))
				else:
					data.append(row[col].value)
			data = tuple(data)
			print (data)
			pointer.execute(sql,data)
			db.commit()

		pointer.close()
		db.close()
		print("[+] INSERTION ENDED")
	except:
		print("connection to database failed")
except:
	print("File not found")
